"""날짜별 분석 — HY202103 폴더의 모든 측정 날짜를 분리해서 CSV + 그림 생성.

`xml_loader.find_all_xmls` 는 같은 (다이, 밴드) 에 대해 최신 측정만 남기지만,
이 스크립트는 모든 날짜의 측정을 보존해서 시간 변화를 추적할 수 있게 함.

출력:
    res/csv/data_by_date.csv         — Date, Wafer, Band, Row, Col, ER, IL, Vpi, is_problematic
    res/figures/by_date_summary.png  — 날짜별 추이 그림 (물리바운드 밖 = 빨강)

실행:
    python3 src/analyze_by_date.py
"""
import sys, os, re, glob
PROGRAM_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(PROGRAM_ROOT, 'src'))

import pandas as pd
import numpy as np
import multiprocessing
import matplotlib; matplotlib.use('Agg')
import matplotlib.pyplot as plt

from xml_loader import load_die, BAND_OF_FILE
from extract_vpi import status_to_reason
from die_extractor import process_die
from outlier_detect import PHYSICAL_BOUNDS
from plot_common import WAFER_BAND_COLOR


DATA_ROOT = os.path.join(PROGRAM_ROOT, 'data', 'HY202103')


def find_all_xmls_with_dates(root_dir):
    """모든 LMZC/LMZO XML 반환 (dedup 없음). 경로 YYYYMMDD_HHMMSS 에서 날짜 추출."""
    items = []
    for tag in BAND_OF_FILE:
        pattern = os.path.join(root_dir, '**', f'*_DCM_{tag}.xml')
        for fp in sorted(glob.glob(pattern, recursive=True)):
            m = re.search(r'(\d{4})(\d{2})(\d{2})_\d{6}', fp)
            if not m:
                continue
            date_str = f'{m.group(1)}-{m.group(2)}-{m.group(3)}'
            items.append((fp, date_str))
    return items


def _process(args):
    """공통 die_extractor.process_die 사용 + Date 키 prepend."""
    xml_path, date_str = args
    row = process_die(xml_path)
    if row is None:
        return None
    # Date 가 맨 앞에 오도록 새 dict 구성
    out = {'Date': date_str}
    out.update(row)
    return out


def _reason(value, lo, hi):
    """Out-of-bound 사유 문자열. 정상이면 빈 문자열."""
    if pd.isna(value):
        return 'missing'
    if value < lo:
        return f'under: {value:.2f} < {lo}'
    if value > hi:
        return f'over: {value:.2f} > {hi}'
    return ''


def _vpi_reason(row):
    """V_π 전용 reason.

    1순위: vpi_status 꼬리표 (extract_vpi 가 명시적으로 내보낸 상태)
    2순위: physical bound 위반 (over/under)
    """
    status = row.get('vpi_status', 'ok')
    dl = row.get('dlam_dV_pm_per_V')
    # extract_vpi 가 망가짐 플래그를 직접 띄운 경우
    if status != 'ok':
        return status_to_reason(status, dlam_dV_pm=dl)
    # 추출은 성공했으나 물리바운드 위반
    lo, hi = PHYSICAL_BOUNDS['Vpi_V']
    v = row['Vpi_V']
    if pd.isna(v):
        return 'missing'
    if v < lo:
        return f'under: {v:.2f} < {lo}'
    if v > hi:
        return f'over: {v:.2f} > {hi}'
    return ''


def _flag_problematic(df):
    """각 항목별 reason 컬럼 + 종합 is_problematic.

    기준: PHYSICAL_BOUNDS (Si MZM 물리 한계).
    z-score 가 아니라 물리 hard limit 을 쓰는 이유:
        - z-score 는 같은 그룹 내 상대비교일 뿐 (그룹 전체가 망가져도 z는 정상)
        - 물리적으로 불가능한 값은 물리바운드만이 잡아낼 수 있음

    V_π 의 경우 slope filter (extract_vpi.MIN_SLOPE_PM_PER_V) 가
    NaN 으로 처리한 케이스를 별도로 표시:
        'broken: |dλ/dV|=0.06 < 10 pm/V (slope filter)'
    """
    # ER, IL: 단순 reason
    for col in ['ER_dB', 'IL_dB']:
        lo, hi = PHYSICAL_BOUNDS[col]
        df[f'reason_{col}'] = df[col].apply(lambda v: _reason(v, lo, hi))
        df[f'out_of_bound_{col}'] = df[f'reason_{col}'] != ''
    # V_π: slope filter 정보까지 포함
    df['reason_Vpi_V'] = df.apply(_vpi_reason, axis=1)
    df['out_of_bound_Vpi_V'] = df['reason_Vpi_V'] != ''
    df['is_problematic'] = (
        df['out_of_bound_ER_dB'] |
        df['out_of_bound_IL_dB'] |
        df['out_of_bound_Vpi_V']
    )
    return df


def export_xlsx(df, path):
    """xlsx 저장 — 문제 셀(reason 비어있지 않음, is_problematic=True)에 빨간 배경."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'data_by_date'

    headers = list(df.columns)
    ws.append(headers)

    # 헤더 스타일
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    red_fill   = PatternFill(start_color='FFC7C7', end_color='FFC7C7', fill_type='solid')
    red_font   = Font(color='B00020', bold=True)
    metric_cols = ['ER_dB', 'IL_dB', 'Vpi_V']

    for _, row in df.iterrows():
        ws.append([row[c] for c in headers])
        excel_row = ws.max_row

        # is_problematic 셀
        if bool(row.get('is_problematic')):
            ci = headers.index('is_problematic') + 1
            cell = ws.cell(row=excel_row, column=ci)
            cell.fill = red_fill
            cell.font = red_font

        # reason_X 가 비어있지 않으면 reason + 원본 metric 둘 다 빨강
        for m in metric_cols:
            rc = f'reason_{m}'
            if rc in headers and row.get(rc):
                # reason 컬럼
                ci = headers.index(rc) + 1
                ws.cell(row=excel_row, column=ci).fill = red_fill
                ws.cell(row=excel_row, column=ci).font = red_font
                # 원본 metric 값
                mi = headers.index(m) + 1
                ws.cell(row=excel_row, column=mi).fill = red_fill
                ws.cell(row=excel_row, column=mi).font = red_font

    # 열 너비 자동 (대략)
    for col_idx, name in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            12, min(28, len(name) + 2))

    wb.save(path)
    print(f'XLSX 저장: {path}')


def analyze(data_root=DATA_ROOT):
    """모든 날짜 측정을 추출해서 DataFrame 반환 (is_problematic 포함)."""
    items = find_all_xmls_with_dates(data_root)
    with multiprocessing.Pool() as pool:
        results = pool.map(_process, items)
    rows = [r for r in results if r is not None]
    df = pd.DataFrame(rows).sort_values(['Date', 'Band', 'Wafer', 'Row', 'Col'])
    df = df.reset_index(drop=True)
    df = _flag_problematic(df)
    return df


def plot_by_date(df, save_path):
    """날짜별 ER/IL/Vpi 추이 — 물리바운드 밖은 빨간색."""
    metrics = [
        ('ER_dB', 'Extinction Ratio (dB)'),
        ('IL_dB', 'Insertion Loss (dB)'),
        ('Vpi_V', 'V_pi (V)'),
    ]
    dates = sorted(df['Date'].unique())
    date_to_x = {d: i for i, d in enumerate(dates)}

    fig, axes = plt.subplots(3, 1, figsize=(11, 12), dpi=140, sharex=True)
    rng = np.random.default_rng(42)

    for ax, (col, label) in zip(axes, metrics):
        lo, hi = PHYSICAL_BOUNDS[col]
        # 물리 신뢰 영역
        ax.axhspan(lo, hi, color='lightgreen', alpha=0.18, zorder=0,
                   label=f'physical bound [{lo}, {hi}]')

        # 정상 값
        ok = ~df[f'out_of_bound_{col}']
        for (w, b), color in WAFER_BAND_COLOR.items():
            sub = df[ok & (df['Wafer'] == w) & (df['Band'] == b)]
            if sub.empty:
                continue
            xs = [date_to_x[d] + rng.normal(0, 0.06) for d in sub['Date']]
            ax.scatter(xs, sub[col], facecolor=color, edgecolor='black',
                       s=40, lw=0.4, alpha=0.85, label=f'{w}[{b}]', zorder=3)

        # 문제 값 → 빨간색 (속 빈 마커)
        bad = df[df[f'out_of_bound_{col}']]
        if not bad.empty:
            xs = [date_to_x[d] + rng.normal(0, 0.06) for d in bad['Date']]
            ax.scatter(xs, bad[col], facecolor='red', edgecolor='darkred',
                       s=80, lw=1.2, marker='x', zorder=5,
                       label=f'⚠ out of bound (n={len(bad)})')

        ax.set_ylabel(label, fontsize=11)
        ax.grid(alpha=0.3, axis='y')

        # Vpi 가 너무 크게 튀면 y축 자르고 위쪽 텍스트로 알림
        if col == 'Vpi_V' and df[col].max() > hi * 3:
            ax.set_ylim(0, hi * 2)
            n_over = int((df[col] > hi * 2).sum())
            if n_over > 0:
                ax.text(0.99, 0.95, f'⚠ {n_over}개 값이 축 위로 벗어남',
                        transform=ax.transAxes, ha='right', va='top',
                        color='red', fontweight='bold', fontsize=10)

        # 범례 — 항목별로 중복 제거
        h, l = ax.get_legend_handles_labels()
        seen, uniq_h, uniq_l = set(), [], []
        for hi_, li in zip(h, l):
            if li not in seen:
                seen.add(li); uniq_h.append(hi_); uniq_l.append(li)
        ax.legend(uniq_h, uniq_l, fontsize=8, loc='best', framealpha=0.85)

    axes[-1].set_xticks(range(len(dates)))
    axes[-1].set_xticklabels(dates, rotation=30, ha='right')
    axes[-1].set_xlabel('Measurement Date')
    fig.suptitle('Per-Date Summary  —  problematic values in red',
                 fontsize=14, fontweight='bold', y=1.00)
    plt.tight_layout()
    plt.savefig(save_path, bbox_inches='tight')
    plt.close(fig)
    print(f'By-date plot saved: {save_path}')


def export_and_plot(df=None, csv_path=None, xlsx_path=None, fig_path=None):
    """run.py 에서 호출용. df 없으면 분석부터 수행."""
    if df is None:
        df = analyze(DATA_ROOT)
    if csv_path is None:
        csv_path = os.path.join(PROGRAM_ROOT, 'res', 'csv', 'data_by_date.csv')
    if xlsx_path is None:
        xlsx_path = os.path.join(PROGRAM_ROOT, 'res', 'csv', 'data_by_date.xlsx')
    if fig_path is None:
        fig_path = os.path.join(PROGRAM_ROOT, 'res', 'figures', 'by_date_summary.png')
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    os.makedirs(os.path.dirname(fig_path), exist_ok=True)
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f'CSV 저장: {csv_path}')
    # all-wafers (alias) + wafer 별 분리
    out_dir = os.path.dirname(csv_path)
    df.to_csv(os.path.join(out_dir, 'data_by_date_all_wafers.csv'),
              index=False, encoding='utf-8-sig')
    for w in sorted(df['Wafer'].unique()):
        sub = df[df['Wafer'] == w]
        sub.to_csv(os.path.join(out_dir, f'data_by_date_{w}.csv'),
                   index=False, encoding='utf-8-sig')
    export_xlsx(df, xlsx_path)
    plot_by_date(df, fig_path)
    return df


def main():
    print('=' * 60)
    print(' 날짜별 분석 시작')
    print('=' * 60)
    items = find_all_xmls_with_dates(DATA_ROOT)
    print(f'\n발견된 다이-측정: {len(items)}개')
    print('병렬 추출 중...')
    df = analyze(DATA_ROOT)
    print(f'        → {len(df)}개 측정 처리 완료')
    n_bad = int(df['is_problematic'].sum())
    print(f'        → 물리바운드 위반: {n_bad}개')

    export_and_plot(df)

    # 날짜별 요약
    print('\n[날짜별 측정 수 (Wafer/Band)]')
    summary = df.groupby(['Date', 'Wafer', 'Band']).size().unstack(
        level=['Wafer', 'Band'], fill_value=0)
    print(summary)

    print('\n[날짜별 중앙값]')
    medians = df.groupby('Date')[['ER_dB', 'IL_dB', 'Vpi_V']].median().round(2)
    print(medians)


if __name__ == '__main__':
    main()
