"""다이별 종합 평가 (evaluation) 모듈 — 모든 측정 (날짜별 보존).

입력은 data_by_date.csv (98 측정).  dedup 안 한 모든 날짜 측정을 한
파일에 통합 평가한다.

출력:
    res/csv/evaluation.csv    데이터 형태
    res/csv/evaluation.xlsx   빨간/노랑 배경 + 비고란 + figures_per_die hyperlink

평가 항목:

    1. Vpi extraction        : vpi_status 가 ok 가 아니면 FAIL
    2. ER  physical bound    : 10 ≤ ER  ≤ 45 dB    벗어나면 FAIL
    3. IL  physical bound    : -15 ≤ IL ≤ -1 dB    벗어나면 FAIL
    4. Vpi physical bound    :   2 ≤ V_π ≤ 60 V    벗어나면 FAIL
    5. ER  Robust Z          : |z| > 3  → WARN  (그룹: Date, Wafer, Band)
    6. IL  Robust Z          : 〃
    7. Vpi Robust Z          : 〃
    8. Linearity R^2         : >= 0.95 PASS / >= 0.90 WARN / >= 0.50 CAUTION / < 0.50 FAIL
    9. Splitter imbalance    : < 0.5 dB PASS / < 1.0 dB OK / >= 1.0 WARN
   10. MZM section loss      : < 1.0 dB PASS / < 2.0 dB OK / >= 2.0 WARN

종합 (overall):
    FAIL  하나라도 있으면 FAIL  (XLSX 행 빨간 배경)
    WARN  하나라도 있고 FAIL 없으면 WARN (노랑 배경)
    그 외 PASS  (흰 배경)

XLSX 의 `figures` 컬럼 = 그 측정의 figures_per_die 폴더로 hyperlink
(정확한 날짜 폴더로).
"""
import sys, os
from urllib.parse import quote as _urlquote
PROGRAM_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(PROGRAM_ROOT, 'src'))

import pandas as pd
import numpy as np

from outlier_detect import PHYSICAL_BOUNDS


CSV_PATH    = os.path.join(PROGRAM_ROOT, 'res', 'csv', 'data_by_date.csv')
FIG_ROOT    = os.path.join(PROGRAM_ROOT, 'res', 'figures_per_die')
OUT_CSV     = os.path.join(PROGRAM_ROOT, 'res', 'csv', 'evaluation.csv')
OUT_XLSX    = os.path.join(PROGRAM_ROOT, 'res', 'csv', 'evaluation.xlsx')

# GitHub repository — figures 컬럼 hyperlink 의 기본 URL.
# 폴더 구조: res/figures_per_die/<date>/<band>-band/<wafer>/(r,c)/
GITHUB_FIG_BASE = 'https://github.com/hanseo-k/auto/tree/main/res/figures_per_die'


# ──────────────────────────────────────────────────────────────────────
# 평가 임계값
# ──────────────────────────────────────────────────────────────────────
LINEARITY_THRESHOLDS = {'pass': 0.95, 'warn': 0.90, 'caution': 0.50}
IMBALANCE_THRESHOLDS = {'pass': 0.5,  'ok':   1.0}
LOSS_THRESHOLDS      = {'pass': 1.0,  'ok':   2.0}


# ──────────────────────────────────────────────────────────────────────
# Robust Z (그룹: Date, Wafer, Band)
# ──────────────────────────────────────────────────────────────────────
def add_robust_z(df):
    """data_by_date.csv 에 robust_z_* 컬럼이 없으므로 직접 계산."""
    for col in ['ER_dB', 'IL_dB', 'Vpi_V']:
        z_col = f'robust_z_{col}'
        df[z_col] = np.nan
        for _, grp in df.groupby(['Date', 'Wafer', 'Band']):
            vals = grp[col].dropna()
            if len(vals) < 3:
                continue
            m = vals.median()
            mad = (vals - m).abs().median()
            sigma = 1.4826 * mad
            if sigma == 0:
                continue
            for idx in grp.index:
                x = grp.loc[idx, col]
                if pd.isna(x):
                    continue
                df.at[idx, z_col] = round((x - m) / sigma, 2)
    return df


# ──────────────────────────────────────────────────────────────────────
# 평가 (다이 한 행)
# ──────────────────────────────────────────────────────────────────────
def _evaluate_row(row):
    issues_fail = []
    issues_warn = []
    cat = {}

    # 1. Vpi extraction status
    vpi_status = str(row.get('vpi_status', 'unknown'))
    if vpi_status != 'ok':
        cat['vpi_extraction'] = 'FAIL'
        issues_fail.append(f'Vpi extraction failed ({vpi_status})')
    else:
        cat['vpi_extraction'] = 'PASS'

    # 2. ER physical
    er = row.get('ER_dB')
    er_lo, er_hi = PHYSICAL_BOUNDS['ER_dB']
    if pd.isna(er):
        cat['er_phys'] = 'FAIL'; issues_fail.append('ER missing')
    elif er < er_lo or er > er_hi:
        cat['er_phys'] = 'FAIL'
        issues_fail.append(f'ER {er:.2f} dB out of bound [{er_lo}, {er_hi}]')
    else:
        cat['er_phys'] = 'PASS'

    # 3. IL physical
    il = row.get('IL_dB')
    il_lo, il_hi = PHYSICAL_BOUNDS['IL_dB']
    if pd.isna(il):
        cat['il_phys'] = 'FAIL'; issues_fail.append('IL missing')
    elif il < il_lo or il > il_hi:
        cat['il_phys'] = 'FAIL'
        issues_fail.append(f'IL {il:.2f} dB out of bound [{il_lo}, {il_hi}]')
    else:
        cat['il_phys'] = 'PASS'

    # 4. Vpi physical
    vpi = row.get('Vpi_V')
    vp_lo, vp_hi = PHYSICAL_BOUNDS['Vpi_V']
    if pd.isna(vpi):
        if cat['vpi_extraction'] == 'PASS':
            cat['vpi_phys'] = 'FAIL'; issues_fail.append('Vpi missing')
        else:
            cat['vpi_phys'] = 'FAIL'
    elif vpi < vp_lo or vpi > vp_hi:
        cat['vpi_phys'] = 'FAIL'
        issues_fail.append(f'Vpi {vpi:.2f} V out of bound [{vp_lo}, {vp_hi}]')
    else:
        cat['vpi_phys'] = 'PASS'

    # 5-7. Robust Z
    for col, name in [('ER_dB', 'er'), ('IL_dB', 'il'), ('Vpi_V', 'vpi')]:
        z = row.get(f'robust_z_{col}')
        if pd.notna(z) and abs(z) > 3:
            cat[f'{name}_z'] = 'WARN'
            issues_warn.append(f'{col} Robust Z = {z:+.2f} (|z|>3)')
        else:
            cat[f'{name}_z'] = 'PASS'

    # 8. Linearity R^2
    r2 = row.get('R2_dlam_vs_V')
    if pd.isna(r2):
        cat['linearity'] = 'FAIL'
        if cat['vpi_extraction'] == 'PASS':
            issues_fail.append('Linearity R² missing')
    elif r2 < LINEARITY_THRESHOLDS['caution']:
        cat['linearity'] = 'FAIL'
        issues_fail.append(f'Linearity R²={r2:.3f} (< 0.5, poor fit)')
    elif r2 < LINEARITY_THRESHOLDS['warn']:
        cat['linearity'] = 'CAUTION'
        issues_warn.append(f'Linearity R²={r2:.3f} (< 0.9, marginal)')
    elif r2 < LINEARITY_THRESHOLDS['pass']:
        cat['linearity'] = 'WARN'
        issues_warn.append(f'Linearity R²={r2:.3f} (< 0.95)')
    else:
        cat['linearity'] = 'PASS'

    # 9. Splitter imbalance
    imb = row.get('imbalance_dB')
    if pd.isna(imb):
        cat['imbalance'] = 'WARN'; issues_warn.append('imbalance missing')
    elif imb >= IMBALANCE_THRESHOLDS['ok']:
        cat['imbalance'] = 'WARN'
        issues_warn.append(f'Imbalance {imb:.2f} dB (>=1.0, MMI imbalance)')
    elif imb >= IMBALANCE_THRESHOLDS['pass']:
        cat['imbalance'] = 'OK'
    else:
        cat['imbalance'] = 'PASS'

    # 10. MZM section loss
    loss = row.get('mzm_loss_dB')
    if pd.isna(loss):
        cat['loss'] = 'WARN'; issues_warn.append('MZM loss missing')
    elif loss >= LOSS_THRESHOLDS['ok']:
        cat['loss'] = 'WARN'
        issues_warn.append(f'MZM loss {loss:.2f} dB (>=2.0, high)')
    elif loss >= LOSS_THRESHOLDS['pass']:
        cat['loss'] = 'OK'
    else:
        cat['loss'] = 'PASS'

    # 종합
    if issues_fail:
        overall = 'FAIL'
    elif issues_warn:
        overall = 'WARN'
    else:
        overall = 'PASS'
    notes = '; '.join(issues_fail + issues_warn)

    out = {'overall': overall, 'notes': notes}
    out.update(cat)
    return out


def _find_figure_dir(date, wafer, band, row, col):
    """정확한 날짜 폴더의 figures_per_die 경로."""
    path = os.path.join(FIG_ROOT, str(date), f'{band}-band', wafer,
                        f'({int(row)},{int(col)})')
    return path if os.path.isdir(path) else ''


def build_evaluation(df):
    df = add_robust_z(df.copy())
    base_cols = ['Date', 'Wafer', 'Band', 'Row', 'Col', 'Width_nm', 'Length_um',
                 'ER_dB', 'IL_dB', 'Vpi_V', 'Vpi_L_V_cm',
                 'R2_dlam_vs_V', 'quality_grade',
                 'imbalance_dB', 'mzm_loss_dB']
    base_cols = [c for c in base_cols if c in df.columns]

    rows = []
    for _, r in df.iterrows():
        ev = _evaluate_row(r)
        rec = {c: r.get(c) for c in base_cols}
        rec.update(ev)
        rec['figures'] = _find_figure_dir(
            r['Date'], r['Wafer'], r['Band'], r['Row'], r['Col'])
        rows.append(rec)

    out_df = pd.DataFrame(rows)
    col_order = (
        base_cols +
        ['overall', 'vpi_extraction', 'er_phys', 'il_phys', 'vpi_phys',
         'er_z', 'il_z', 'vpi_z', 'linearity', 'imbalance', 'loss',
         'notes', 'figures']
    )
    col_order = [c for c in col_order if c in out_df.columns]
    return out_df[col_order]


# ──────────────────────────────────────────────────────────────────────
# XLSX (빨간/노랑 배경 + hyperlink)
# ──────────────────────────────────────────────────────────────────────
def write_xlsx(df, path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = 'evaluation'

    headers = list(df.columns)
    ws.append(headers)
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9',
                              fill_type='solid')
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.font = Font(bold=True); c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    fail_fill = PatternFill(start_color='FFC7C7', end_color='FFC7C7',
                            fill_type='solid')
    warn_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC',
                            fill_type='solid')
    fail_font = Font(color='B00020', bold=True)
    warn_font = Font(color='996600', bold=False)

    fig_col_idx = headers.index('figures') + 1 if 'figures' in headers else None

    for _, row in df.iterrows():
        ws.append([row[c] for c in headers])
        excel_row = ws.max_row
        overall = row.get('overall')

        if overall == 'FAIL':
            row_fill, row_font = fail_fill, fail_font
        elif overall == 'WARN':
            row_fill, row_font = warn_fill, warn_font
        else:
            row_fill, row_font = None, None

        if row_fill is not None:
            for ci in range(1, len(headers) + 1):
                cell = ws.cell(row=excel_row, column=ci)
                cell.fill = row_fill
                if row_font is not None:
                    cell.font = row_font

        if fig_col_idx is not None and row.get('figures'):
            cell = ws.cell(row=excel_row, column=fig_col_idx)
            cell.value = 'open'
            # GitHub web URL — 다른 사람과 공유해도 클릭 즉시 web view 로 이동.
            # 폴더명에 괄호/콤마 들어가니 path segment 별로 안전 인코딩.
            rel = os.path.relpath(row['figures'], FIG_ROOT)
            url_path = '/'.join(_urlquote(seg, safe='') for seg in rel.split(os.sep))
            cell.hyperlink = f'{GITHUB_FIG_BASE}/{url_path}'
            cell.font = Font(color='0000EE', underline='single',
                             bold=(overall == 'FAIL'))

    widths = {
        'Date': 12, 'Wafer': 8, 'Band': 6, 'Row': 6, 'Col': 6,
        'Width_nm': 9, 'Length_um': 10,
        'ER_dB': 9, 'IL_dB': 9, 'Vpi_V': 9, 'Vpi_L_V_cm': 11,
        'R2_dlam_vs_V': 12, 'quality_grade': 8,
        'imbalance_dB': 12, 'mzm_loss_dB': 12,
        'overall': 9, 'vpi_extraction': 14,
        'er_phys': 8, 'il_phys': 8, 'vpi_phys': 9,
        'er_z': 6, 'il_z': 6, 'vpi_z': 7,
        'linearity': 10, 'imbalance': 11, 'loss': 7,
        'notes': 70, 'figures': 9,
    }
    for ci, name in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, 12)
    ws.freeze_panes = 'A2'
    wb.save(path)


def main():
    df = pd.read_csv(CSV_PATH)
    print('=' * 70)
    print(f' 평가 (evaluation) — 측정 {len(df)} 개 (모든 날짜)')
    print('=' * 70)

    eval_df = build_evaluation(df)
    eval_df.to_csv(OUT_CSV, index=False, encoding='utf-8-sig')
    write_xlsx(eval_df, OUT_XLSX)

    counts = eval_df['overall'].value_counts()
    print('\n[종합 분포]')
    for k in ('PASS', 'WARN', 'FAIL'):
        print(f'  {k:5s}: {counts.get(k, 0):3d} 측정')

    print('\n[날짜별 종합]')
    for date, grp in eval_df.groupby('Date'):
        n = len(grp)
        f = (grp['overall'] == 'FAIL').sum()
        w = (grp['overall'] == 'WARN').sum()
        p = (grp['overall'] == 'PASS').sum()
        print(f'  {date}:  PASS {p}/{n}, WARN {w}/{n}, FAIL {f}/{n}')

    fails = eval_df[eval_df['overall'] == 'FAIL']
    if not fails.empty:
        print(f'\n[FAIL {len(fails)}개 — 앞 5개]')
        for _, r in fails.head(5).iterrows():
            print(f'  {r["Date"]} {r["Wafer"]} {r["Band"]} '
                  f'({int(r["Row"])},{int(r["Col"])}): {r["notes"]}')

    print(f'\nCSV  저장: {OUT_CSV}')
    print(f'XLSX 저장: {OUT_XLSX}')


if __name__ == '__main__':
    main()
